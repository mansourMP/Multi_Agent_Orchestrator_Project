import csv
import json
import os
from pathlib import Path
from typing import Any, Dict

from orion_local_worker_utils import split_items


def parse_positive_int(value: Any, default: int = 0) -> int:
    try:
        parsed = int(value)
    except Exception:
        return default
    return parsed if parsed >= 0 else default


def _spreadsheet_root_path() -> Path:
    configured = str(os.getenv("ORION_SPREADSHEET_ROOT") or "").strip()
    base = Path(configured).expanduser() if configured else Path.cwd()
    return base.resolve()


def _resolve_spreadsheet_target(raw_path: Any) -> tuple[Path, Path]:
    value = str(raw_path or "").strip()
    if not value:
        raise RuntimeError("Spreadsheet file path is required.")
    root = _spreadsheet_root_path()
    candidate = Path(value).expanduser()
    target = (root / candidate).resolve() if not candidate.is_absolute() else candidate.resolve()
    try:
        target.relative_to(root)
    except Exception as exc:
        raise RuntimeError(f"Spreadsheet path must stay inside: {root}") from exc
    if target.suffix.lower() not in {".csv", ".xlsx"}:
        raise RuntimeError("Spreadsheet path must end with .csv or .xlsx")
    target.parent.mkdir(parents=True, exist_ok=True)
    return target, root


def _parse_json_safe(raw: Any) -> Any:
    if isinstance(raw, (dict, list)):
        return raw
    if not isinstance(raw, str):
        return None
    text = raw.strip()
    if not text:
        return None
    try:
        return json.loads(text)
    except Exception:
        return None


def _parse_key_value_text(raw: str) -> Dict[str, str]:
    result: Dict[str, str] = {}
    for line in str(raw or "").replace("\r", "\n").split("\n"):
        cleaned = line.strip()
        if not cleaned:
            continue
        separator = ":" if ":" in cleaned else "=" if "=" in cleaned else None
        if not separator:
            continue
        key, value = cleaned.split(separator, 1)
        key_clean = key.strip().lower()
        value_clean = value.strip()
        if key_clean and value_clean:
            result[key_clean] = value_clean
    return result


def _normalize_row_payload(payload: Any) -> list[Dict[str, Any]]:
    rows: list[Dict[str, Any]] = []
    if isinstance(payload, list):
        for idx, item in enumerate(payload):
            if isinstance(item, dict):
                rows.append({str(k): item[k] for k in item.keys()})
            elif isinstance(item, list):
                rows.append({f"col_{col_idx + 1}": value for col_idx, value in enumerate(item)})
            elif item is not None:
                rows.append({"value": item, "row_index": idx})
    elif isinstance(payload, dict):
        nested = payload.get("rows")
        if isinstance(nested, list):
            rows.extend(_normalize_row_payload(nested))
        elif payload:
            rows.append({str(k): payload[k] for k in payload.keys()})
    return rows


def _read_csv_preview(path: Path, row_limit: int) -> tuple[list[str], list[Dict[str, Any]], int]:
    if not path.exists():
        raise RuntimeError(f"Spreadsheet file not found: {path}")
    preview: list[Dict[str, Any]] = []
    rows_read = 0
    with path.open("r", newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        columns = [str(col or "").strip() for col in (reader.fieldnames or []) if str(col or "").strip()]
        for row in reader:
            rows_read += 1
            if len(preview) < row_limit:
                preview.append({str(k): row.get(k) for k in row.keys()})
    return columns, preview, rows_read


def _read_xlsx_preview(path: Path, sheet_name: str, row_limit: int) -> tuple[list[str], list[Dict[str, Any]], int]:
    if not path.exists():
        raise RuntimeError(f"Spreadsheet file not found: {path}")
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError("openpyxl is required for .xlsx operations.") from exc
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    rows_raw = list(ws.iter_rows(values_only=True))
    if not rows_raw:
        return [], [], 0
    header_raw = rows_raw[0]
    columns = [
        str(value).strip() if str(value or "").strip() else f"col_{idx + 1}"
        for idx, value in enumerate(header_raw)
    ]
    preview: list[Dict[str, Any]] = []
    rows_read = max(0, len(rows_raw) - 1)
    for row in rows_raw[1 : 1 + row_limit]:
        preview.append({columns[idx]: row[idx] if idx < len(row) else None for idx in range(len(columns))})
    return columns, preview, rows_read


def _normalize_action_id(raw: Any) -> str:
    return str(raw or "").strip().lower().replace(" ", "_").replace("-", "_")


def _parse_spreadsheet_pack_inputs(pack_inputs: Dict[str, Any]) -> Dict[str, Any]:
    file_path = str(
        pack_inputs.get("file_path")
        or pack_inputs.get("path")
        or pack_inputs.get("inbox")
        or ""
    ).strip()
    raw_operation = pack_inputs.get("operation") or pack_inputs.get("leads") or ""
    raw_payload = pack_inputs.get("payload") or pack_inputs.get("values") or pack_inputs.get("slots") or ""

    operation_cfg = _parse_json_safe(raw_operation)
    if not isinstance(operation_cfg, dict):
        operation_cfg = {}
    if not operation_cfg:
        kv = _parse_key_value_text(str(raw_operation or ""))
        if kv:
            operation_cfg = {**operation_cfg, **kv}

    payload_cfg = _parse_json_safe(raw_payload)
    if payload_cfg is None:
        payload_cfg = {}

    operation = _normalize_action_id(
        operation_cfg.get("operation")
        or operation_cfg.get("op")
        or pack_inputs.get("operation_type")
        or raw_operation
    )
    if operation in {"", "sheet"}:
        operation = "read"
    if operation == "edit":
        operation = "update"
    if operation not in {"read", "append", "update", "create"}:
        operation = "read"

    sheet_name = str(
        operation_cfg.get("sheet_name")
        or operation_cfg.get("sheet")
        or pack_inputs.get("sheet_name")
        or "Sheet1"
    ).strip() or "Sheet1"
    row_limit = parse_positive_int(operation_cfg.get("row_limit"), 50)
    row_limit = min(max(1, row_limit), 500)
    row_index = parse_positive_int(
        operation_cfg.get("row_index")
        if isinstance(operation_cfg, dict)
        else pack_inputs.get("row_index"),
        0,
    )
    overwrite = bool(operation_cfg.get("overwrite")) if isinstance(operation_cfg, dict) else False

    rows = _normalize_row_payload(payload_cfg)
    if not rows and isinstance(raw_payload, str):
        fallback_lines = split_items(raw_payload)
        rows = [{"value": line} for line in fallback_lines]

    update_values: Dict[str, Any] = {}
    if isinstance(payload_cfg, dict):
        candidate = payload_cfg.get("values")
        if isinstance(candidate, dict):
            update_values = {str(k): candidate[k] for k in candidate.keys()}
        elif operation == "update":
            update_values = {
                str(k): payload_cfg[k]
                for k in payload_cfg.keys()
                if str(k) not in {"rows", "row_index", "sheet_name", "operation", "op"}
            }
    if isinstance(payload_cfg, dict) and "row_index" in payload_cfg:
        row_index = parse_positive_int(payload_cfg.get("row_index"), row_index)
    if isinstance(payload_cfg, dict) and "overwrite" in payload_cfg:
        overwrite = bool(payload_cfg.get("overwrite"))
    if isinstance(payload_cfg, dict) and "sheet_name" in payload_cfg and str(payload_cfg.get("sheet_name")).strip():
        sheet_name = str(payload_cfg.get("sheet_name")).strip()

    return {
        "file_path": file_path,
        "operation": operation,
        "sheet_name": sheet_name,
        "row_limit": row_limit,
        "row_index": row_index,
        "overwrite": overwrite,
        "rows": rows,
        "update_values": update_values,
    }


def build_spreadsheet_pack_result(pack_inputs: Dict[str, Any]) -> tuple[str, Dict[str, Any]]:
    parsed = _parse_spreadsheet_pack_inputs(pack_inputs)
    target_path, root_path = _resolve_spreadsheet_target(parsed.get("file_path"))
    operation = str(parsed.get("operation") or "read")
    sheet_name = str(parsed.get("sheet_name") or "Sheet1")
    row_limit = parse_positive_int(parsed.get("row_limit"), 50)
    row_index = parse_positive_int(parsed.get("row_index"), 0)
    overwrite = bool(parsed.get("overwrite"))
    rows: list[Dict[str, Any]] = parsed.get("rows") if isinstance(parsed.get("rows"), list) else []
    update_values: Dict[str, Any] = parsed.get("update_values") if isinstance(parsed.get("update_values"), dict) else {}

    format_name = target_path.suffix.lower().replace(".", "")
    columns: list[str] = []
    preview: list[Dict[str, Any]] = []
    rows_read = 0
    rows_written = 0
    action_payload: Dict[str, Any] = {
        "file_path": str(target_path.relative_to(root_path)),
        "sheet_name": sheet_name,
    }

    if format_name == "csv":
        if operation == "read":
            columns, preview, rows_read = _read_csv_preview(target_path, row_limit)
            action_payload["action"] = "spreadsheet_read"
            action_payload["row_limit"] = row_limit
        elif operation == "create":
            if target_path.exists() and not overwrite:
                raise RuntimeError(f"File already exists: {target_path.relative_to(root_path)} (set overwrite=true).")
            rows_to_write = list(rows)
            columns = sorted({key for row in rows_to_write for key in row.keys()}) if rows_to_write else ["value"]
            with target_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(handle, fieldnames=columns)
                writer.writeheader()
                for row in rows_to_write:
                    writer.writerow({key: row.get(key, "") for key in columns})
            rows_written = len(rows_to_write)
            columns, preview, rows_read = _read_csv_preview(target_path, row_limit)
            action_payload["action"] = "spreadsheet_create"
            action_payload["rows"] = rows_to_write[:20]
            action_payload["overwrite"] = overwrite
        elif operation == "append":
            if not rows:
                raise RuntimeError("Append operation requires payload rows.")
            existing_columns: list[str] = []
            existing_rows: list[Dict[str, Any]] = []
            if target_path.exists():
                with target_path.open("r", newline="", encoding="utf-8") as handle:
                    reader = csv.DictReader(handle)
                    existing_columns = [str(col or "").strip() for col in (reader.fieldnames or []) if str(col or "").strip()]
                    for row in reader:
                        existing_rows.append({str(k): row.get(k) for k in row.keys()})
            if not existing_columns:
                existing_columns = sorted({key for row in rows for key in row.keys()})
            if not existing_columns:
                existing_columns = ["value"]
            for row in rows:
                extra = [key for key in row.keys() if key not in existing_columns]
                if extra:
                    raise RuntimeError(f"Append payload has unknown columns: {', '.join(extra)}.")
            with target_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(handle, fieldnames=existing_columns)
                writer.writeheader()
                for row in existing_rows:
                    writer.writerow({key: row.get(key, "") for key in existing_columns})
                for row in rows:
                    writer.writerow({key: row.get(key, "") for key in existing_columns})
            rows_written = len(rows)
            columns, preview, rows_read = _read_csv_preview(target_path, row_limit)
            action_payload["action"] = "spreadsheet_append"
            action_payload["rows"] = rows[:20]
        else:  # update
            if not target_path.exists():
                raise RuntimeError(f"Spreadsheet file not found: {target_path.relative_to(root_path)}")
            if not update_values:
                raise RuntimeError("Update operation requires payload.values object.")
            with target_path.open("r", newline="", encoding="utf-8") as handle:
                reader = csv.DictReader(handle)
                existing_columns = [str(col or "").strip() for col in (reader.fieldnames or []) if str(col or "").strip()]
                data_rows = [{str(k): row.get(k) for k in row.keys()} for row in reader]
            if row_index >= len(data_rows):
                raise RuntimeError(f"row_index {row_index} out of range (rows={len(data_rows)}).")
            for key in update_values.keys():
                if str(key) not in existing_columns:
                    existing_columns.append(str(key))
            target = data_rows[row_index]
            for key, value in update_values.items():
                target[str(key)] = value
            with target_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(handle, fieldnames=existing_columns)
                writer.writeheader()
                for row in data_rows:
                    writer.writerow({key: row.get(key, "") for key in existing_columns})
            rows_written = 1
            columns, preview, rows_read = _read_csv_preview(target_path, row_limit)
            action_payload["action"] = "spreadsheet_update"
            action_payload["row_index"] = row_index
            action_payload["values"] = {str(k): update_values[k] for k in update_values.keys()}
    else:
        try:
            from openpyxl import Workbook, load_workbook  # type: ignore
        except Exception as exc:
            raise RuntimeError("openpyxl is required for .xlsx operations.") from exc
        if operation == "read":
            columns, preview, rows_read = _read_xlsx_preview(target_path, sheet_name, row_limit)
            action_payload["action"] = "spreadsheet_read"
            action_payload["row_limit"] = row_limit
        else:
            if target_path.exists():
                wb = load_workbook(target_path)
            else:
                wb = Workbook()
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active
                ws.title = sheet_name[:31]

            if operation == "create":
                if target_path.exists() and not overwrite:
                    raise RuntimeError(f"File already exists: {target_path.relative_to(root_path)} (set overwrite=true).")
                wb = Workbook()
                ws = wb.active
                ws.title = sheet_name[:31]
                rows_to_write = list(rows)
                write_columns = sorted({key for row in rows_to_write for key in row.keys()}) if rows_to_write else ["value"]
                ws.append(write_columns)
                for row in rows_to_write:
                    ws.append([row.get(column) for column in write_columns])
                wb.save(target_path)
                rows_written = len(rows_to_write)
                columns, preview, rows_read = _read_xlsx_preview(target_path, sheet_name, row_limit)
                action_payload["action"] = "spreadsheet_create"
                action_payload["rows"] = rows_to_write[:20]
                action_payload["overwrite"] = overwrite
            elif operation == "append":
                if not rows:
                    raise RuntimeError("Append operation requires payload rows.")
                if ws.max_row < 1:
                    header = sorted({key for row in rows for key in row.keys()}) or ["value"]
                    ws.append(header)
                header = [
                    str(ws.cell(row=1, column=idx).value).strip() if str(ws.cell(row=1, column=idx).value or "").strip() else f"col_{idx}"
                    for idx in range(1, ws.max_column + 1)
                ]
                for row in rows:
                    extra = [key for key in row.keys() if key not in header]
                    for key in extra:
                        header.append(key)
                        ws.cell(row=1, column=len(header), value=key)
                    ws.append([row.get(col) for col in header])
                wb.save(target_path)
                rows_written = len(rows)
                columns, preview, rows_read = _read_xlsx_preview(target_path, sheet_name, row_limit)
                action_payload["action"] = "spreadsheet_append"
                action_payload["rows"] = rows[:20]
            else:  # update
                if ws.max_row < 2:
                    raise RuntimeError("Update requires an existing data row.")
                if not update_values:
                    raise RuntimeError("Update operation requires payload.values object.")
                excel_row = row_index + 2
                if excel_row > ws.max_row:
                    raise RuntimeError(f"row_index {row_index} out of range (rows={max(0, ws.max_row - 1)}).")
                header = [
                    str(ws.cell(row=1, column=idx).value).strip() if str(ws.cell(row=1, column=idx).value or "").strip() else f"col_{idx}"
                    for idx in range(1, ws.max_column + 1)
                ]
                for key, value in update_values.items():
                    key_text = str(key)
                    if key_text not in header:
                        header.append(key_text)
                        ws.cell(row=1, column=len(header), value=key_text)
                    col_idx = header.index(key_text) + 1
                    ws.cell(row=excel_row, column=col_idx, value=value)
                wb.save(target_path)
                rows_written = 1
                columns, preview, rows_read = _read_xlsx_preview(target_path, sheet_name, row_limit)
                action_payload["action"] = "spreadsheet_update"
                action_payload["row_index"] = row_index
                action_payload["values"] = {str(k): update_values[k] for k in update_values.keys()}

    tool_action = str(action_payload.get("action") or "spreadsheet_read")
    summary = (
        f"Spreadsheet Ops completed: {tool_action.replace('spreadsheet_', '')} on "
        f"{action_payload.get('file_path')} ({format_name}). "
        f"Rows read: {rows_read}. Rows written: {rows_written}."
    )
    data = {
        "pack_id": "spreadsheet-ops-v1",
        "summary": summary,
        "inputs": {
            "file_path": str(action_payload.get("file_path") or ""),
            "operation": operation,
            "sheet_name": sheet_name,
            "rows_input": len(rows),
        },
        "outputs": {
            "operation": operation,
            "file_path": str(action_payload.get("file_path") or ""),
            "sheet_name": sheet_name,
            "format": format_name,
            "columns": columns,
            "preview": preview,
            "rows_read": rows_read,
            "rows_written": rows_written,
            "actions": [action_payload],
            "outbound_actions": rows_written,
            "urgent_count": 0,
        },
        "next_steps": [
            "Review preview rows for correctness.",
            "Use append/update for additional changes when needed.",
            "Share or export the updated spreadsheet.",
        ],
    }
    return summary, data
