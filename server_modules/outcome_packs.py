from __future__ import annotations

import csv
import hashlib
import json
import mimetypes
import os
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from server_modules.connector_metadata import _connector_identity_signature
from server_modules.customer_ops_pack import execute_customer_ops_pack
from server_modules.external_write_safety import execute_external_write_once, stable_value_fingerprint
from server_modules.microsoft_365_graph import (
    microsoft_365_download_drive_file,
    microsoft_365_normalize_drive_path,
    microsoft_365_upload_drive_file,
)
from server_modules.office_ooxml import (
    DOCX_MIME,
    PPTX_MIME,
    build_docx,
    build_pptx,
    build_updated_docx,
    build_updated_pptx,
    normalize_deck_slides,
    normalize_doc_sections,
)
from server_modules.runtime_policy import (
    COMPETITOR_BRIEF_PACK_ID,
    CUSTOMER_OPS_PACK_ID,
    DOCUMENT_STUDIO_PACK_ID,
    LOCAL_EXECUTION_PACK_ID,
    SPREADSHEET_OPS_PACK_ID,
    WEEKLY_CONTENT_PACK_ID,
    normalize_action_id,
    parse_positive_int,
    parse_text_items,
)
from server_modules.vault_helpers import resolve_vault_credential
from server_modules.vault_store import _openssl_decrypt, load_vault


_CHANNEL_ALIASES = {
    "ig": "Instagram",
    "instagram": "Instagram",
    "fb": "Facebook",
    "facebook": "Facebook",
    "x": "X",
    "twitter": "X",
    "linkedin": "LinkedIn",
    "li": "LinkedIn",
    "tiktok": "TikTok",
    "tt": "TikTok",
    "yt": "YouTube",
    "youtube": "YouTube",
    "pinterest": "Pinterest",
    "pin": "Pinterest",
    "snap": "Snapchat",
    "snapchat": "Snapchat",
    "email": "Email",
    "newsletter": "Email",
    "whatsapp": "WhatsApp",
    "sms": "SMS",
    "text": "SMS",
}


def _connector_account_identity(
    connector_provider: str,
    connector_credential_id: Any,
    connector_credentials: Dict[str, Any],
) -> Optional[str]:
    credential_token = str(connector_credential_id or "").strip()
    if credential_token:
        return f"credential:{credential_token}"
    signature = _connector_identity_signature(
        connector_provider,
        connector_credentials if isinstance(connector_credentials, dict) else {},
    ) or ""
    signature = signature.strip()
    if signature:
        return f"signature:{stable_value_fingerprint(signature)}"
    return None


def normalize_channel_name(raw: Any) -> str:
    text = str(raw or "").strip()
    if not text:
        return "Unknown"
    key = text.lower()
    if key in _CHANNEL_ALIASES:
        return _CHANNEL_ALIASES[key]
    if len(text) <= 3 and text.isalpha():
        return text.upper()
    return text.title()


def channel_format(channel: str) -> str:
    key = str(channel or "").strip().lower()
    if key in {"email", "newsletter"}:
        return "email"
    if key in {"sms", "text", "whatsapp"}:
        return "message"
    if key in {"instagram", "facebook", "linkedin", "tiktok", "x", "twitter", "youtube", "pinterest", "snapchat"}:
        return "social_post"
    return "post"


def execute_weekly_content_pack(context: Dict[str, Any], run_id: Optional[str] = None) -> Dict[str, Any]:
    metadata = context.get("metadata") if isinstance(context.get("metadata"), dict) else {}
    pack_inputs = metadata.get("pack_inputs") if isinstance(metadata.get("pack_inputs"), dict) else {}

    topics = parse_text_items(pack_inputs.get("topics") or pack_inputs.get("inbox") or "", max_items=20)
    channels_raw = parse_text_items(pack_inputs.get("channels") or pack_inputs.get("leads") or "", max_items=20)
    offers = parse_text_items(pack_inputs.get("offers") or pack_inputs.get("slots") or "", max_items=20)

    if not topics:
        topics = [
            "Before/after service outcomes",
            "Customer FAQ answered clearly",
            "Seasonal promotion highlight",
            "Local trust story",
            "Team expertise spotlight",
        ]
    if not channels_raw:
        channels_raw = ["Instagram", "Facebook", "Email"]
    if not offers:
        offers = ["Book this week for priority scheduling", "Reply to get a quick quote"]

    channels = [normalize_channel_name(item) for item in channels_raw]
    days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    item_count = min(7, max(5, len(topics)))

    plan_items: List[Dict[str, Any]] = []
    for idx in range(item_count):
        topic = topics[idx % len(topics)]
        channel = channels[idx % len(channels)]
        offer = offers[idx % len(offers)]
        output = {
            "id": f"content-{idx+1}",
            "day": days[idx % len(days)],
            "channel": channel,
            "format": channel_format(channel),
            "topic": topic,
            "headline": f"{topic} — {offer}",
            "cta": offer,
            "status": "draft_ready",
        }
        plan_items.append(output)

    summary = (
        f"Weekly Content Studio completed: drafted {len(plan_items)} post briefs across "
        f"{len(set(channels))} channels with clear CTA recommendations."
    )

    return {
        "pack_id": WEEKLY_CONTENT_PACK_ID,
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "summary": summary,
        "inputs": {
            "topics_count": len(topics),
            "channels_count": len(channels),
            "offers_count": len(offers),
        },
        "outputs": {
            "content_plan": plan_items,
            "outbound_actions": len(plan_items),
            "urgent_count": 0,
        },
        "next_steps": [
            "Approve the top 3 posts and publish this week.",
            "Assign one owner per channel for final edits.",
            "Review engagement after 7 days and iterate.",
        ],
    }


def competitor_threat_level(raw: str) -> str:
    text = raw.lower()
    high_tokens = ["cheap", "low price", "same day", "24/7", "free", "ai", "automation"]
    medium_tokens = ["nearby", "local", "bundle", "promo", "offer"]
    if any(token in text for token in high_tokens):
        return "high"
    if any(token in text for token in medium_tokens):
        return "medium"
    return "low"


def execute_competitor_brief_pack(context: Dict[str, Any], run_id: Optional[str] = None) -> Dict[str, Any]:
    metadata = context.get("metadata") if isinstance(context.get("metadata"), dict) else {}
    pack_inputs = metadata.get("pack_inputs") if isinstance(metadata.get("pack_inputs"), dict) else {}

    competitors = parse_text_items(pack_inputs.get("competitors") or pack_inputs.get("inbox") or "", max_items=12)
    positioning = parse_text_items(pack_inputs.get("positioning") or pack_inputs.get("leads") or "", max_items=6)
    objectives = parse_text_items(pack_inputs.get("objectives") or pack_inputs.get("slots") or "", max_items=6)

    if not competitors:
        competitors = [
            "Competitor A | strong reviews | discount pricing",
            "Competitor B | heavy social ads | fast booking",
            "Competitor C | premium pricing | niche audience",
        ]
    if not positioning:
        positioning = ["Reliable service quality", "Fast response", "Clear pricing"]
    if not objectives:
        objectives = ["Increase qualified leads", "Improve close rate", "Protect margin"]

    briefs: List[Dict[str, Any]] = []
    for idx, competitor in enumerate(competitors):
        threat = competitor_threat_level(competitor)
        action = (
            "Counter with proof-based testimonials and faster response SLA."
            if threat == "high"
            else "Differentiate with clearer offer packaging."
            if threat == "medium"
            else "Monitor monthly and keep baseline response."
        )
        briefs.append(
            {
                "id": f"competitor-{idx+1}",
                "name": competitor.split("|", 1)[0].strip() or f"Competitor {idx + 1}",
                "notes": competitor,
                "threat_level": threat,
                "recommended_play": action,
            }
        )

    high_threat = len([item for item in briefs if item["threat_level"] == "high"])
    summary = (
        f"Competitor Brief Digest completed: analyzed {len(briefs)} competitors. "
        f"High-threat competitors identified: {high_threat}."
    )

    return {
        "pack_id": COMPETITOR_BRIEF_PACK_ID,
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "summary": summary,
        "inputs": {
            "competitors_count": len(competitors),
            "positioning_points": len(positioning),
            "objectives_count": len(objectives),
        },
        "outputs": {
            "briefs": briefs,
            "high_threat_count": high_threat,
            "outbound_actions": 0,
            "urgent_count": high_threat,
        },
        "next_steps": [
            "Prioritize one response play for each high-threat competitor.",
            "Update your top offer and landing copy based on this brief.",
            "Re-run this digest next week to track movement.",
        ],
    }


def _resolve_vault_credential(credential_id: str, workspace_id: Optional[str]) -> Dict[str, Any]:
    return resolve_vault_credential(load_vault, _openssl_decrypt, credential_id, workspace_id)


def _spreadsheet_root_path() -> Path:
    configured = str(os.getenv("ORION_SPREADSHEET_ROOT") or "").strip()
    base = Path(configured).expanduser() if configured else Path.cwd()
    return base.resolve()


def _document_root_path() -> Path:
    configured = str(os.getenv("ORION_DOCUMENT_ROOT") or os.getenv("ORION_SPREADSHEET_ROOT") or "").strip()
    base = Path(configured).expanduser() if configured else Path.cwd()
    return base.resolve()


def _resolve_spreadsheet_target(raw_path: Any) -> tuple[Path, Path]:
    value = str(raw_path or "").strip()
    if not value:
        raise RuntimeError("Spreadsheet file path is required.")
    root = _spreadsheet_root_path()
    candidate = Path(value).expanduser()
    target = (root / candidate).resolve() if not candidate.is_absolute() else candidate.resolve()
    try:
        target.relative_to(root)
    except Exception as exc:
        raise RuntimeError(f"Spreadsheet path must stay inside: {root}") from exc
    if target.suffix.lower() not in {".csv", ".xlsx"}:
        raise RuntimeError("Spreadsheet path must end with .csv or .xlsx")
    target.parent.mkdir(parents=True, exist_ok=True)
    return target, root


def _resolve_document_target(raw_path: Any) -> tuple[Path, Path]:
    value = str(raw_path or "").strip()
    if not value:
        raise RuntimeError("Document file path is required.")
    root = _document_root_path()
    candidate = Path(value).expanduser()
    target = (root / candidate).resolve() if not candidate.is_absolute() else candidate.resolve()
    try:
        target.relative_to(root)
    except Exception as exc:
        raise RuntimeError(f"Document path must stay inside: {root}") from exc
    if target.suffix.lower() not in {".docx", ".pptx"}:
        raise RuntimeError("Document path must end with .docx or .pptx")
    target.parent.mkdir(parents=True, exist_ok=True)
    return target, root


def _parse_json_safe(raw: Any) -> Any:
    if isinstance(raw, (dict, list)):
        return raw
    if not isinstance(raw, str):
        return None
    text = raw.strip()
    if not text:
        return None
    try:
        return json.loads(text)
    except Exception:
        return None


def _parse_key_value_text(raw: str) -> Dict[str, str]:
    result: Dict[str, str] = {}
    for line in str(raw or "").replace("\r", "\n").split("\n"):
        cleaned = line.strip()
        if not cleaned:
            continue
        separator = ":" if ":" in cleaned else "=" if "=" in cleaned else None
        if not separator:
            continue
        key, value = cleaned.split(separator, 1)
        key_clean = key.strip().lower()
        value_clean = value.strip()
        if key_clean and value_clean:
            result[key_clean] = value_clean
    return result


def _normalize_row_payload(payload: Any) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    if isinstance(payload, list):
        for idx, item in enumerate(payload):
            if isinstance(item, dict):
                rows.append({str(k): item[k] for k in item.keys()})
            elif isinstance(item, list):
                rows.append({f"col_{col_idx + 1}": value for col_idx, value in enumerate(item)})
            elif item is not None:
                rows.append({"value": item, "row_index": idx})
    elif isinstance(payload, dict):
        nested = payload.get("rows")
        if isinstance(nested, list):
            rows.extend(_normalize_row_payload(nested))
        elif payload:
            rows.append({str(k): payload[k] for k in payload.keys()})
    return rows


def _read_csv_preview(path: Path, row_limit: int) -> tuple[List[str], List[Dict[str, Any]], int]:
    if not path.exists():
        raise RuntimeError(f"Spreadsheet file not found: {path}")
    preview: List[Dict[str, Any]] = []
    rows_read = 0
    with path.open("r", newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        columns = [str(col or "").strip() for col in (reader.fieldnames or []) if str(col or "").strip()]
        for row in reader:
            rows_read += 1
            if len(preview) < row_limit:
                preview.append({str(k): row.get(k) for k in row.keys()})
    return columns, preview, rows_read


def _read_xlsx_preview(path: Path, sheet_name: str, row_limit: int) -> tuple[List[str], List[Dict[str, Any]], int]:
    if not path.exists():
        raise RuntimeError(f"Spreadsheet file not found: {path}")
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError("openpyxl is required for .xlsx operations.") from exc

    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    rows_raw = list(ws.iter_rows(values_only=True))
    if not rows_raw:
        return [], [], 0
    header_raw = rows_raw[0]
    columns = [
        str(value).strip() if str(value or "").strip() else f"col_{idx + 1}"
        for idx, value in enumerate(header_raw)
    ]
    preview: List[Dict[str, Any]] = []
    rows_read = max(0, len(rows_raw) - 1)
    for row in rows_raw[1 : 1 + row_limit]:
        preview.append({columns[idx]: row[idx] if idx < len(row) else None for idx in range(len(columns))})
    return columns, preview, rows_read


def _parse_spreadsheet_pack_inputs(pack_inputs: Dict[str, Any]) -> Dict[str, Any]:
    file_path = str(
        pack_inputs.get("file_path")
        or pack_inputs.get("path")
        or pack_inputs.get("inbox")
        or ""
    ).strip()
    raw_operation = pack_inputs.get("operation") or pack_inputs.get("leads") or ""
    raw_payload = pack_inputs.get("payload") or pack_inputs.get("values") or pack_inputs.get("slots") or ""

    operation_cfg = _parse_json_safe(raw_operation)
    if not isinstance(operation_cfg, dict):
        operation_cfg = {}
    if not operation_cfg:
        kv = _parse_key_value_text(str(raw_operation or ""))
        if kv:
            operation_cfg = {**operation_cfg, **kv}

    payload_cfg = _parse_json_safe(raw_payload)
    if payload_cfg is None:
        payload_cfg = {}

    operation = normalize_action_id(
        operation_cfg.get("operation")
        or operation_cfg.get("op")
        or pack_inputs.get("operation_type")
        or raw_operation
    )
    if operation in {"", "sheet"}:
        operation = "read"
    if operation == "edit":
        operation = "update"
    if operation not in {"read", "append", "update", "create"}:
        operation = "read"

    sheet_name = str(
        operation_cfg.get("sheet_name")
        or operation_cfg.get("sheet")
        or pack_inputs.get("sheet_name")
        or "Sheet1"
    ).strip() or "Sheet1"

    row_limit = parse_positive_int(
        operation_cfg.get("row_limit")
        or operation_cfg.get("limit")
        or pack_inputs.get("row_limit")
        or 50,
        50,
    )
    row_limit = max(1, min(row_limit, 200))

    row_index = parse_positive_int(
        operation_cfg.get("row_index")
        or operation_cfg.get("row")
        or pack_inputs.get("row_index")
        or 0,
        0,
    )

    overwrite = bool(operation_cfg.get("overwrite")) if isinstance(operation_cfg, dict) else False

    update_values: Dict[str, Any] = {}
    rows: List[Dict[str, Any]] = []
    if isinstance(payload_cfg, dict) and "values" in payload_cfg:
        if isinstance(payload_cfg.get("values"), dict):
            update_values = dict(payload_cfg.get("values"))
    if isinstance(payload_cfg, dict) and "rows" in payload_cfg:
        rows = _normalize_row_payload(payload_cfg.get("rows"))
    elif payload_cfg:
        rows = _normalize_row_payload(payload_cfg)

    if isinstance(payload_cfg, dict) and "row_index" in payload_cfg:
        row_index = parse_positive_int(payload_cfg.get("row_index"), row_index)
    if isinstance(payload_cfg, dict) and "overwrite" in payload_cfg:
        overwrite = bool(payload_cfg.get("overwrite"))
    if isinstance(payload_cfg, dict) and "sheet_name" in payload_cfg and str(payload_cfg.get("sheet_name")).strip():
        sheet_name = str(payload_cfg.get("sheet_name")).strip()

    return {
        "file_path": file_path,
        "operation": operation,
        "sheet_name": sheet_name,
        "row_limit": row_limit,
        "row_index": row_index,
        "overwrite": overwrite,
        "rows": rows,
        "update_values": update_values,
    }


def execute_spreadsheet_ops_pack(context: Dict[str, Any], run_id: Optional[str] = None) -> Dict[str, Any]:
    metadata = context.get("metadata") if isinstance(context.get("metadata"), dict) else {}
    pack_inputs = metadata.get("pack_inputs") if isinstance(metadata.get("pack_inputs"), dict) else {}
    parsed = _parse_spreadsheet_pack_inputs(pack_inputs)

    workspace_id = context.get("workspace_id") or metadata.get("workspace_id")
    connector_credential_id = metadata.get("connector_credential_id")
    connector_payload = metadata.get("pack_connectors") if isinstance(metadata.get("pack_connectors"), dict) else {}
    connector_provider = ""
    connector_credentials: Dict[str, Any] = {}

    if isinstance(connector_credential_id, str) and connector_credential_id.strip():
        try:
            secret = _resolve_vault_credential(connector_credential_id.strip(), str(workspace_id) if workspace_id else None)
            provider = str(secret.get("_provider") or "").strip()
            if provider == "microsoft_365":
                connector_provider = provider
                connector_credentials = dict(secret)
        except Exception:
            pass
    elif isinstance(connector_payload, dict):
        payload_provider = str(connector_payload.get("connector") or connector_payload.get("_provider") or "").strip()
        if payload_provider == "microsoft_365":
            connector_provider = payload_provider
            connector_credentials = dict(connector_payload)
    connector_account_id = _connector_account_identity(
        connector_provider,
        connector_credential_id,
        connector_credentials,
    )

    source_path = str(parsed.get("file_path") or "").strip()
    remote_storage = source_path.lower().startswith(("onedrive:/", "m365:/", "onedrive://", "m365://"))
    remote_drive_path = microsoft_365_normalize_drive_path(source_path) if remote_storage else ""
    if remote_storage and connector_provider != "microsoft_365":
        raise RuntimeError("OneDrive spreadsheet paths require a connected Microsoft 365 connector.")

    operation = str(parsed.get("operation") or "read")
    sheet_name = str(parsed.get("sheet_name") or "Sheet1")
    row_limit = parse_positive_int(parsed.get("row_limit"), 50)
    row_index = parse_positive_int(parsed.get("row_index"), 0)
    overwrite = bool(parsed.get("overwrite"))
    rows: List[Dict[str, Any]] = parsed.get("rows") if isinstance(parsed.get("rows"), list) else []
    update_values: Dict[str, Any] = parsed.get("update_values") if isinstance(parsed.get("update_values"), dict) else {}

    sync_dir = Path(".orion-m365-sync")
    sync_dir.mkdir(parents=True, exist_ok=True)

    remote_exists = False
    if remote_storage:
        file_name = Path(remote_drive_path).name or f"spreadsheet-{uuid.uuid4().hex}.xlsx"
        target_path = sync_dir / f"{uuid.uuid4().hex}-{file_name}"
        root_path = sync_dir
        try:
            target_path.write_bytes(microsoft_365_download_drive_file(connector_credentials, remote_drive_path))
            remote_exists = True
        except Exception as exc:
            detail = str(exc).lower()
            if operation == "create" and ("http 404" in detail or "not found" in detail or "itemnotfound" in detail):
                remote_exists = False
            else:
                raise RuntimeError(f"OneDrive file unavailable: {exc}") from exc
    else:
        target_path, root_path = _resolve_spreadsheet_target(parsed.get("file_path"))

    display_path = f"onedrive:/{remote_drive_path}" if remote_storage else str(target_path.relative_to(root_path))

    def run_local_spreadsheet_ops(local_target_path: Path, display_file_path: str) -> Dict[str, Any]:
        format_name = local_target_path.suffix.lower().replace(".", "")
        if format_name not in {"csv", "xlsx"}:
            raise RuntimeError("Spreadsheet path must end in .csv or .xlsx.")

        columns: List[str] = []
        preview: List[Dict[str, Any]] = []
        rows_read = 0
        rows_written = 0
        action_payload: Dict[str, Any] = {
            "file_path": display_file_path,
            "sheet_name": sheet_name,
            "storage": "onedrive" if remote_storage else "local",
            "connector": connector_provider or None,
        }

        if format_name == "csv":
            if operation == "read":
                columns, preview, rows_read = _read_csv_preview(local_target_path, row_limit)
                action_payload["action"] = "spreadsheet_read"
                action_payload["row_limit"] = row_limit
            elif operation == "create":
                if local_target_path.exists() and not overwrite:
                    raise RuntimeError(f"File already exists: {display_file_path} (set overwrite=true).")
                rows_to_write = list(rows)
                columns = sorted({key for row in rows_to_write for key in row.keys()}) if rows_to_write else ["value"]
                with local_target_path.open("w", newline="", encoding="utf-8") as handle:
                    writer = csv.DictWriter(handle, fieldnames=columns)
                    writer.writeheader()
                    for row in rows_to_write:
                        writer.writerow({key: row.get(key, "") for key in columns})
                rows_written = len(rows_to_write)
                columns, preview, rows_read = _read_csv_preview(local_target_path, row_limit)
                action_payload["action"] = "spreadsheet_create"
                action_payload["rows"] = rows_to_write[:20]
                action_payload["overwrite"] = overwrite
            elif operation == "append":
                if not rows:
                    raise RuntimeError("Append operation requires payload rows.")
                existing_columns: List[str] = []
                existing_rows: List[Dict[str, Any]] = []
                if local_target_path.exists():
                    with local_target_path.open("r", newline="", encoding="utf-8") as handle:
                        reader = csv.DictReader(handle)
                        existing_columns = [str(col or "").strip() for col in (reader.fieldnames or []) if str(col or "").strip()]
                        for row in reader:
                            existing_rows.append({str(k): row.get(k) for k in row.keys()})
                if not existing_columns:
                    existing_columns = sorted({key for row in rows for key in row.keys()})
                if not existing_columns:
                    existing_columns = ["value"]
                for row in rows:
                    extra = [key for key in row.keys() if key not in existing_columns]
                    if extra:
                        raise RuntimeError(f"Append payload has unknown columns: {', '.join(extra)}.")
                with local_target_path.open("w", newline="", encoding="utf-8") as handle:
                    writer = csv.DictWriter(handle, fieldnames=existing_columns)
                    writer.writeheader()
                    for row in existing_rows:
                        writer.writerow({key: row.get(key, "") for key in existing_columns})
                    for row in rows:
                        writer.writerow({key: row.get(key, "") for key in existing_columns})
                rows_written = len(rows)
                columns, preview, rows_read = _read_csv_preview(local_target_path, row_limit)
                action_payload["action"] = "spreadsheet_append"
                action_payload["rows"] = rows[:20]
            else:
                if not local_target_path.exists():
                    raise RuntimeError(f"Spreadsheet file not found: {display_file_path}")
                if not update_values:
                    raise RuntimeError("Update operation requires payload.values object.")
                with local_target_path.open("r", newline="", encoding="utf-8") as handle:
                    reader = csv.DictReader(handle)
                    existing_columns = [str(col or "").strip() for col in (reader.fieldnames or []) if str(col or "").strip()]
                    data_rows = [{str(k): row.get(k) for k in row.keys()} for row in reader]
                if row_index >= len(data_rows):
                    raise RuntimeError(f"row_index {row_index} out of range (rows={len(data_rows)}).")
                for key in update_values.keys():
                    if str(key) not in existing_columns:
                        existing_columns.append(str(key))
                target = data_rows[row_index]
                for key, value in update_values.items():
                    target[str(key)] = value
                with local_target_path.open("w", newline="", encoding="utf-8") as handle:
                    writer = csv.DictWriter(handle, fieldnames=existing_columns)
                    writer.writeheader()
                    for row in data_rows:
                        writer.writerow({key: row.get(key, "") for key in existing_columns})
                rows_written = 1
                columns, preview, rows_read = _read_csv_preview(local_target_path, row_limit)
                action_payload["action"] = "spreadsheet_update"
                action_payload["row_index"] = row_index
                action_payload["values"] = {str(k): update_values[k] for k in update_values.keys()}
        else:
            try:
                from openpyxl import Workbook, load_workbook  # type: ignore
            except Exception as exc:
                raise RuntimeError("openpyxl is required for .xlsx operations.") from exc

            if operation == "read":
                columns, preview, rows_read = _read_xlsx_preview(local_target_path, sheet_name, row_limit)
                action_payload["action"] = "spreadsheet_read"
                action_payload["row_limit"] = row_limit
            else:
                if local_target_path.exists():
                    wb = load_workbook(local_target_path)
                else:
                    wb = Workbook()
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.active
                    ws.title = sheet_name[:31]

                if operation == "create":
                    if local_target_path.exists() and not overwrite:
                        raise RuntimeError(f"File already exists: {display_file_path} (set overwrite=true).")
                    wb = Workbook()
                    ws = wb.active
                    ws.title = sheet_name[:31]
                    rows_to_write = list(rows)
                    write_columns = sorted({key for row in rows_to_write for key in row.keys()}) if rows_to_write else ["value"]
                    ws.append(write_columns)
                    for row in rows_to_write:
                        ws.append([row.get(column) for column in write_columns])
                    wb.save(local_target_path)
                    rows_written = len(rows_to_write)
                    columns, preview, rows_read = _read_xlsx_preview(local_target_path, sheet_name, row_limit)
                    action_payload["action"] = "spreadsheet_create"
                    action_payload["rows"] = rows_to_write[:20]
                    action_payload["overwrite"] = overwrite
                elif operation == "append":
                    if not rows:
                        raise RuntimeError("Append operation requires payload rows.")
                    if ws.max_row < 1:
                        header = sorted({key for row in rows for key in row.keys()}) or ["value"]
                        ws.append(header)
                    header = [
                        str(ws.cell(row=1, column=idx).value).strip() if str(ws.cell(row=1, column=idx).value or "").strip() else f"col_{idx}"
                        for idx in range(1, ws.max_column + 1)
                    ]
                    for row in rows:
                        extra = [key for key in row.keys() if key not in header]
                        for key in extra:
                            header.append(key)
                            ws.cell(row=1, column=len(header), value=key)
                        ws.append([row.get(col) for col in header])
                    wb.save(local_target_path)
                    rows_written = len(rows)
                    columns, preview, rows_read = _read_xlsx_preview(local_target_path, sheet_name, row_limit)
                    action_payload["action"] = "spreadsheet_append"
                    action_payload["rows"] = rows[:20]
                else:
                    if ws.max_row < 2:
                        raise RuntimeError("Update requires an existing data row.")
                    if not update_values:
                        raise RuntimeError("Update operation requires payload.values object.")
                    excel_row = row_index + 2
                    if excel_row > ws.max_row:
                        raise RuntimeError(f"row_index {row_index} out of range (rows={max(0, ws.max_row - 1)}).")
                    header = [
                        str(ws.cell(row=1, column=idx).value).strip() if str(ws.cell(row=1, column=idx).value or "").strip() else f"col_{idx}"
                        for idx in range(1, ws.max_column + 1)
                    ]
                    for key, value in update_values.items():
                        key_text = str(key)
                        if key_text not in header:
                            header.append(key_text)
                            ws.cell(row=1, column=len(header), value=key_text)
                        col_idx = header.index(key_text) + 1
                        ws.cell(row=excel_row, column=col_idx, value=value)
                    wb.save(local_target_path)
                    rows_written = 1
                    columns, preview, rows_read = _read_xlsx_preview(local_target_path, sheet_name, row_limit)
                    action_payload["action"] = "spreadsheet_update"
                    action_payload["row_index"] = row_index
                    action_payload["values"] = {str(k): update_values[k] for k in update_values.keys()}

        tool_action = str(action_payload.get("action") or "spreadsheet_read")
        return {
            "summary": (
                f"Spreadsheet Ops completed: {tool_action.replace('spreadsheet_', '')} on "
                f"{display_file_path} ({format_name}). "
                f"Rows read: {rows_read}. Rows written: {rows_written}."
            ),
            "inputs": {
                "file_path": display_file_path,
                "operation": operation,
                "sheet_name": sheet_name,
                "rows_input": len(rows),
            },
            "outputs": {
                "operation": operation,
                "file_path": display_file_path,
                "sheet_name": sheet_name,
                "format": format_name,
                "columns": columns,
                "preview": preview,
                "rows_read": rows_read,
                "rows_written": rows_written,
                "actions": [action_payload],
                "outbound_actions": rows_written,
                "urgent_count": 0,
                "storage": "onedrive" if remote_storage else "local",
                "connector": connector_provider or None,
                "remote_path": display_file_path if remote_storage else None,
            },
        }

    if remote_storage and not remote_exists and operation != "create":
        raise RuntimeError(f"OneDrive file not found: {display_path}")

    result = run_local_spreadsheet_ops(target_path, display_path)

    if remote_storage and operation != "read":
        content_type = mimetypes.guess_type(str(target_path.name))[0] or "application/octet-stream"
        content_bytes = target_path.read_bytes()
        upload_result = execute_external_write_once(
            run_id=run_id,
            context=context,
            step_key=f"spreadsheet_ops:onedrive_sync:{display_file_path}:{operation}",
            category="spreadsheet_ops_onedrive_sync",
            operation={"action_id": "upload_drive_file", "path": remote_drive_path},
            target_system="microsoft_365",
            account_identity=connector_account_id,
            target_identity={"action_id": "upload_drive_file", "path": remote_drive_path},
            payload={"path": remote_drive_path, "content_type": content_type, "content_sha256": hashlib.sha256(content_bytes).hexdigest()},
            execute=lambda: microsoft_365_upload_drive_file(
                connector_credentials,
                remote_drive_path,
                content_bytes,
                content_type=content_type,
            ),
        )
        duplicate_guard = upload_result.get("duplicate_guard") if isinstance(upload_result, dict) else {}
        duplicate_protected = isinstance(duplicate_guard, dict) and not bool(duplicate_guard.get("executed", True))
        if isinstance(result.get("outputs"), dict):
            actions = result["outputs"].get("actions")
            if isinstance(actions, list):
                result["outputs"]["actions"] = [
                    {**item, "executed": not duplicate_protected, "duplicate_protected": duplicate_protected}
                    if isinstance(item, dict) else item
                    for item in actions
                ]
        result["summary"] = (
            result["summary"].rstrip(".") + " Duplicate-protected OneDrive sync reused the prior upload."
            if duplicate_protected
            else result["summary"].rstrip(".") + " Synced to OneDrive."
        )

    next_steps = [
        "Review preview rows for correctness.",
        "Use append/update with approval for additional changes.",
        "Export or share the updated spreadsheet when ready.",
    ]
    if remote_storage:
        next_steps[-1] = "Share or continue editing the synced OneDrive workbook when ready."

    return {
        "pack_id": SPREADSHEET_OPS_PACK_ID,
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "summary": result["summary"],
        "inputs": result["inputs"],
        "outputs": result["outputs"],
        "next_steps": next_steps,
    }


def _parse_document_studio_pack_inputs(pack_inputs: Dict[str, Any]) -> Dict[str, Any]:
    file_path = str(pack_inputs.get("file_path") or pack_inputs.get("path") or pack_inputs.get("inbox") or "").strip()
    raw_operation = pack_inputs.get("operation") or pack_inputs.get("leads") or ""
    raw_payload = pack_inputs.get("payload") or pack_inputs.get("content") or pack_inputs.get("slots") or pack_inputs.get("values") or ""

    operation_cfg = _parse_json_safe(raw_operation)
    if not isinstance(operation_cfg, dict):
        operation_cfg = {}
    if not operation_cfg:
        kv = _parse_key_value_text(str(raw_operation or ""))
        if kv:
            operation_cfg = {**operation_cfg, **kv}

    payload_cfg = _parse_json_safe(raw_payload)
    payload_value: Any = payload_cfg if payload_cfg is not None else raw_payload

    operation = normalize_action_id(
        operation_cfg.get("operation")
        or operation_cfg.get("op")
        or pack_inputs.get("operation_type")
        or raw_operation
        or "create"
    )
    if operation in {"", "new", "write"}:
        operation = "create"
    if operation in {"edit", "append"}:
        operation = "update"
    if operation not in {"create", "update"}:
        operation = "create"

    overwrite = bool(operation_cfg.get("overwrite")) if isinstance(operation_cfg, dict) else False
    title_hint = str(
        operation_cfg.get("title")
        or (payload_cfg.get("title") if isinstance(payload_cfg, dict) else "")
        or Path(file_path).stem.replace("-", " ").replace("_", " ").strip().title()
        or "Empyralis Document"
    ).strip() or "Empyralis Document"

    return {
        "file_path": file_path,
        "operation": operation,
        "overwrite": overwrite,
        "title": title_hint,
        "payload": payload_value,
    }


def execute_document_studio_pack(context: Dict[str, Any], run_id: Optional[str] = None) -> Dict[str, Any]:
    metadata = context.get("metadata") if isinstance(context.get("metadata"), dict) else {}
    pack_inputs = metadata.get("pack_inputs") if isinstance(metadata.get("pack_inputs"), dict) else {}
    parsed = _parse_document_studio_pack_inputs(pack_inputs)

    workspace_id = context.get("workspace_id") or metadata.get("workspace_id")
    connector_credential_id = metadata.get("connector_credential_id")
    connector_payload = metadata.get("pack_connectors") if isinstance(metadata.get("pack_connectors"), dict) else {}
    connector_provider = ""
    connector_credentials: Dict[str, Any] = {}

    if isinstance(connector_credential_id, str) and connector_credential_id.strip():
        try:
            secret = _resolve_vault_credential(connector_credential_id.strip(), str(workspace_id) if workspace_id else None)
            provider = str(secret.get("_provider") or "").strip()
            if provider == "microsoft_365":
                connector_provider = provider
                connector_credentials = dict(secret)
        except Exception:
            pass
    elif isinstance(connector_payload, dict):
        payload_provider = str(connector_payload.get("connector") or connector_payload.get("_provider") or "").strip()
        if payload_provider == "microsoft_365":
            connector_provider = payload_provider
            connector_credentials = dict(connector_payload)
    connector_account_id = _connector_account_identity(
        connector_provider,
        connector_credential_id,
        connector_credentials,
    )

    source_path = str(parsed.get("file_path") or "").strip()
    remote_storage = source_path.lower().startswith(("onedrive:/", "m365:/", "onedrive://", "m365://"))
    remote_drive_path = microsoft_365_normalize_drive_path(source_path) if remote_storage else ""
    if remote_storage and connector_provider != "microsoft_365":
        raise RuntimeError("OneDrive document paths require a connected Microsoft 365 connector.")

    operation = str(parsed.get("operation") or "create")
    overwrite = bool(parsed.get("overwrite"))
    title_hint = str(parsed.get("title") or "Empyralis Document")

    sync_dir = Path(".orion-m365-sync")
    sync_dir.mkdir(parents=True, exist_ok=True)

    remote_exists = False
    existing_blob: bytes | None = None
    if remote_storage:
        file_name = Path(remote_drive_path).name or f"document-{uuid.uuid4().hex}.docx"
        target_path = sync_dir / f"{uuid.uuid4().hex}-{file_name}"
        root_path = sync_dir
        try:
            existing_blob = microsoft_365_download_drive_file(connector_credentials, remote_drive_path)
            target_path.write_bytes(existing_blob)
            remote_exists = True
        except Exception as exc:
            detail = str(exc).lower()
            if operation == "create" and ("http 404" in detail or "not found" in detail or "itemnotfound" in detail):
                remote_exists = False
                existing_blob = None
            else:
                raise RuntimeError(f"OneDrive file unavailable: {exc}") from exc
    else:
        target_path, root_path = _resolve_document_target(source_path)
        if target_path.exists():
            existing_blob = target_path.read_bytes()

    display_path = f"onedrive:/{remote_drive_path}" if remote_storage else str(target_path.relative_to(root_path))
    file_exists = remote_exists if remote_storage else target_path.exists()
    ext = target_path.suffix.lower()
    if operation == "create" and file_exists and not overwrite:
        raise RuntimeError("Target document already exists. Use update mode or set overwrite=true.")
    if operation == "update" and not file_exists:
        raise RuntimeError("Target document does not exist yet. Use create mode first.")

    preview: List[str] = []
    items_written = 0
    action_payload: Dict[str, Any]
    output_bytes: bytes
    format_name = ext.replace(".", "")

    if ext == ".docx":
        normalized = normalize_doc_sections(parsed.get("payload"), fallback_title=title_hint)
        title = str(normalized.get("title") or title_hint)
        sections = normalized.get("sections") if isinstance(normalized.get("sections"), list) else []
        items_written = len(sections)
        preview = [getattr(section, "heading", "") for section in sections[:4] if str(getattr(section, "heading", "")).strip()]
        output_bytes = build_updated_docx(existing_blob, title, sections) if operation == "update" else build_docx(title, sections)
        action_payload = {
            "action": "document_update" if operation == "update" else "document_create",
            "file_path": display_path,
            "title": title,
            "sections": [{"heading": getattr(section, "heading", ""), "body": list(getattr(section, "body", []))} for section in sections],
            "storage": "onedrive" if remote_storage else "local",
            "connector": connector_provider or None,
            "overwrite": overwrite,
        }
    elif ext == ".pptx":
        normalized = normalize_deck_slides(parsed.get("payload"), fallback_title=title_hint)
        title = str(normalized.get("title") or title_hint)
        slides = normalized.get("slides") if isinstance(normalized.get("slides"), list) else []
        items_written = len(slides)
        preview = [getattr(slide, "title", "") for slide in slides[:4] if str(getattr(slide, "title", "")).strip()]
        output_bytes = build_updated_pptx(existing_blob, title, slides) if operation == "update" else build_pptx(title, slides)
        action_payload = {
            "action": "presentation_update" if operation == "update" else "presentation_create",
            "file_path": display_path,
            "title": title,
            "slides": [{"title": getattr(slide, "title", ""), "bullets": list(getattr(slide, "bullets", []))} for slide in slides],
            "storage": "onedrive" if remote_storage else "local",
            "connector": connector_provider or None,
            "overwrite": overwrite,
        }
    else:
        raise RuntimeError("Document Studio supports only .docx and .pptx targets.")

    target_path.write_bytes(output_bytes)
    if remote_storage:
        content_type = DOCX_MIME if ext == ".docx" else PPTX_MIME
        upload_result = execute_external_write_once(
            run_id=run_id,
            context=context,
            step_key=f"document_studio:onedrive_sync:{display_path}:{operation}",
            category="document_studio_onedrive_sync",
            operation={"action_id": "upload_drive_file", "path": remote_drive_path},
            target_system="microsoft_365",
            account_identity=connector_account_id,
            target_identity={"action_id": "upload_drive_file", "path": remote_drive_path},
            payload={"path": remote_drive_path, "content_type": content_type, "content_sha256": hashlib.sha256(output_bytes).hexdigest()},
            execute=lambda: microsoft_365_upload_drive_file(
                connector_credentials,
                remote_drive_path,
                output_bytes,
                content_type=content_type,
            ),
        )
        duplicate_guard = upload_result.get("duplicate_guard") if isinstance(upload_result, dict) else {}
        duplicate_protected = isinstance(duplicate_guard, dict) and not bool(duplicate_guard.get("executed", True))
        action_payload["executed"] = not duplicate_protected
        action_payload["duplicate_protected"] = duplicate_protected
    else:
        duplicate_protected = False

    storage_label = "OneDrive" if remote_storage else "local"
    kind_label = "Word document" if ext == ".docx" else "PowerPoint deck"
    summary = (
        f"Document Studio completed: {operation}d {kind_label} at {display_path} via {storage_label}. "
        "Duplicate-protected OneDrive sync reused the prior upload."
        if remote_storage and duplicate_protected
        else f"Document Studio completed: {operation}d {kind_label} at {display_path} via {storage_label}."
    )

    return {
        "pack_id": DOCUMENT_STUDIO_PACK_ID,
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "summary": summary,
        "inputs": {
            "file_path": display_path,
            "operation": operation,
            "title": title_hint,
            "storage": "onedrive" if remote_storage else "local",
        },
        "outputs": {
            "operation": operation,
            "file_path": display_path,
            "format": format_name,
            "title": action_payload.get("title") or title_hint,
            "preview": preview,
            "items_written": items_written,
            "actions": [action_payload],
            "outbound_actions": 1,
            "urgent_count": 0,
            "storage": "onedrive" if remote_storage else "local",
            "connector": connector_provider or None,
            "remote_path": display_path if remote_storage else None,
        },
        "next_steps": [
            f"Open the generated {kind_label.lower()} and verify the structure.",
            "Use update mode when you want to append more material instead of recreating the file.",
            "Keep the exact local or OneDrive path for the next revision pass.",
        ],
    }


def execute_outcome_pack(pack_id: str, context: Dict[str, Any], run_id: Optional[str] = None) -> Dict[str, Any]:
    if pack_id == COMPETITOR_BRIEF_PACK_ID:
        return execute_competitor_brief_pack(context, run_id=run_id)
    if pack_id == CUSTOMER_OPS_PACK_ID:
        return execute_customer_ops_pack(context, run_id=run_id)
    if pack_id == WEEKLY_CONTENT_PACK_ID:
        return execute_weekly_content_pack(context, run_id=run_id)
    if pack_id == SPREADSHEET_OPS_PACK_ID:
        return execute_spreadsheet_ops_pack(context, run_id=run_id)
    if pack_id == DOCUMENT_STUDIO_PACK_ID:
        return execute_document_studio_pack(context, run_id=run_id)
    if pack_id == LOCAL_EXECUTION_PACK_ID:
        raise RuntimeError("This local execution pack requires Gateway and cannot run in cloud mode.")
    raise RuntimeError(f"Unsupported outcome pack '{pack_id}'")
